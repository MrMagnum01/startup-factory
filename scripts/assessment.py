#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Startup Factory - Phase 1 Market Assessment engine (single source of truth).
100 scored ideas -> top 50 -> all rankings -> 27-field detail.
Emits: docs/ASSESSMENT.md, docs/scoring.xlsx, data/ideas_all100.json, data/ideas_top50.json
Scores 1-10. Fast Revenue Score = RS*.30 + PP*.25 + EB*.20 + AS*.15 + LC*.10
EB=Ease of Build (high=easy). LC=Low Competition (high=less competition).
Reported "Launch difficulty"=11-EB. "Competition level"=11-LC.
"""
import json, os

CAT = {
 "productized service": dict(
   stack="Astro static landing + Tailwind; Tally/Calendly intake; no DB",
   storage="None - intake to Tally/Formspree inbox + your email",
   pay="Stripe Payment Link or Lemon Squeezy checkout",
   mvp="One-page offer: outcome headline, 3 proof points, fixed scope, price, FAQ, Stripe link + booking form",
   skip="Accounts, dashboard, automated fulfillment, multi-tier pricing, custom backend",
   first10="DM/cold-email 30-50 ideal-fit businesses with a free mini-teardown; post the offer in 2 niche communities",
   test="Send 50 personalised outreach touches; target 1+ paid order or 3+ booked calls in 14 days",
   risk="Fulfillment time caps scale; low cold-reply rates; you trade hours for money until productised further",
   kill="0 paid orders AND <2 booked calls after 100 outreach touches in 14 days"),
 "template pack": dict(
   stack="Astro static product page + Tailwind; asset = downloadable file",
   storage="None - file delivered & licensed via Gumroad/Lemon Squeezy",
   pay="Gumroad / Lemon Squeezy / Stripe Payment Link",
   mvp="Product page: hero preview, what's-inside list, 3 screenshots, price, buy button, refund note, FAQ",
   skip="Membership, drip updates, affiliate program, bundles, accounts",
   first10="Post a free sample in 3 relevant subreddits/Discords; list on Gumroad Discover; 1 short demo clip",
   test="Drive 300-500 targeted visits; target 5+ sales in 14 days at list price",
   risk="Commoditised; piracy & refunds; needs constant distribution to keep selling",
   kill="<1% conversion after 500 targeted visits, or <2 sales in 14 days"),
 "calculator": dict(
   stack="Astro static + vanilla-JS island (client-side math); Tailwind",
   storage="None - 100% client-side",
   pay="Affiliate links + optional 'Pro' Gumroad unlock or lead form",
   mvp="The calculator (correct math), result explainer, 1 affiliate/lead CTA, SEO title/desc, shareable result",
   skip="Accounts, saved history, server APIs, paywalls beyond one upsell",
   first10="Rank for 1 long-tail query; answer 5 relevant Reddit/Quora threads with the tool; 1 Product Hunt post",
   test="Reach 200+ organic/community visits; target 30+ tool completions & 3+ CTA clicks in 14 days",
   risk="SEO is slow & commoditised; thin monetisation per visit unless niche has buyer intent",
   kill="<50 visits AND <1 CTA click after 14 days of posting + indexing"),
 "micro-tool": dict(
   stack="Astro static + client-side JS (optionally WASM); Tailwind",
   storage="None - files processed in-browser, never uploaded",
   pay="Freemium 'Pro' unlock (Gumroad/LemonSqueezy) or affiliate/ads",
   mvp="Working tool, privacy note ('runs in your browser'), 1 upsell CTA, SEO metadata, share button",
   skip="Accounts, cloud storage, batch/queue infra, team features",
   first10="SEO long-tail + answer 5 forum threads; submit to tool directories; 1 launch post",
   test="200+ visits; target 25+ successful uses & 2+ Pro clicks in 14 days",
   risk="Commodity space, hard to rank; free alternatives everywhere",
   kill="<50 visits AND 0 Pro interest after 14 days"),
 "directory": dict(
   stack="Astro static, data from JSON/Markdown; client-side search/filter; Tailwind",
   storage="JSON/Markdown files in repo (no DB)",
   pay="Paid/featured listings + affiliate links + newsletter sponsor slot",
   mvp="Seeded list (40-100 entries), search + 2 filters, detail rows, 'submit listing' form, 1 sponsor slot",
   skip="User accounts, reviews, dynamic submissions DB, claim flows",
   first10="Seed great entries, then tell each listed brand they're featured (warm outreach for upgrades)",
   test="Index + 200 visits; target 1 paid/featured listing or 3 affiliate clicks in 14 days",
   risk="Cold-start (needs both content and traffic); affiliate payouts are thin early",
   kill="<150 visits AND 0 listing/affiliate revenue interest after 14 days"),
 "lead-gen site": dict(
   stack="Astro static landing + Tailwind; Tally/Formspree form; no DB",
   storage="None - leads to form inbox / Google Sheet via webhook",
   pay="Sell qualified leads to local businesses, or affiliate/referral fee per lead",
   mvp="Niche landing, trust signals, short quote form, thank-you page, 1 follow-up email template",
   skip="Quoting engine, accounts, CRM, payment until a buyer is lined up",
   first10="Line up ONE buyer business first; then drive niche traffic (local SEO + 1 paid test) to generate leads",
   test="Generate 5+ qualified leads in 14 days AND get 1 business to agree a per-lead price",
   risk="Two-sided cold start; lead quality disputes; ad costs can exceed lead value",
   kill="<3 leads in 14 days OR no buyer will commit to a per-lead price"),
 "generator": dict(
   stack="Astro static + client-side JS (rules/word-lists/templates, no paid API); Tailwind",
   storage="None - generation is client-side from bundled data",
   pay="Freemium Pro pack (more templates) via Gumroad + affiliate CTA",
   mvp="Generator that returns useful output, copy button, 1 upsell, SEO metadata, examples",
   skip="LLM/paid APIs, accounts, saved projects, server rendering",
   first10="SEO long-tail + community answers; submit to tool directories; 1 launch post",
   test="200+ visits; target 25+ generations & 2+ Pro clicks in 14 days",
   risk="Quality ceiling without an LLM; easy to clone; thin monetisation",
   kill="<50 visits AND 0 Pro interest after 14 days"),
 "landing page": dict(
   stack="Astro static + Tailwind; Tally/Buttondown waitlist",
   storage="None - emails to Buttondown/MailerLite",
   pay="Pre-sell (deposit/early-bird via Stripe) or waitlist -> later paid",
   mvp="Clear promise, who-it's-for, 3 benefits, social proof placeholder, email capture or pre-order",
   skip="The actual product until demand is proven; accounts; backend",
   first10="Post the promise in 3 niche communities; 10 warm DMs; 1 Product Hunt 'coming soon'",
   test="150+ visits; target 30+ waitlist emails or 3 pre-orders in 14 days",
   risk="Waitlist signups != revenue; intent can be soft; you still have to build it",
   kill="<10 waitlist emails AND 0 pre-orders after 150 visits"),
 "mini SaaS": dict(
   stack="Astro/Next + client-side core; Supabase/Neon only if state needed; Tailwind",
   storage="Start stateless; add Supabase/Turso only when a feature requires it",
   pay="Stripe subscription or one-time license (Lemon Squeezy handles tax)",
   mvp="One job-to-be-done done well, single workflow, Stripe paywall on the 1 valuable action",
   skip="Teams, roles, settings sprawl, integrations, mobile app",
   first10="Find 10 people with the exact pain in communities; offer lifetime/early deal for feedback",
   test="10 activated users; target 3 paying or 10 trials in 14 days",
   risk="Scope creep, support load, churn; slowest of the categories to revenue",
   kill="<3 activated users OR 0 willingness-to-pay signals in 14 days"),
}

# id,name,type,desc,target,pain,why_pay,monetization,price,build,channel, rs,pp,eb,as,lc,seo
RAW = [
 (1,"LocalRank Audit","productized service","A fixed-scope local-SEO audit + action plan for service businesses.","Dentists, plumbers, law firms, clinics","They rank below competitors on Google Maps and don't know why","A clear, prioritized fix list that wins them more calls","Stripe Payment Link, done-for-you audit","$149 flat","4-6h","Cold email + local FB groups",9,7,8,6,6,6),
 (2,"ShopSpeed Fix","productized service","Shopify store speed optimization that lifts conversions.","Shopify DTC owners doing $5k-100k/mo","Slow store = lost sales & worse ad ROAS","Faster store directly lifts revenue; clear before/after","Stripe Payment Link, fixed package","$249","5-8h","Cold email + Shopify subreddits",9,8,7,6,6,5),
 (3,"ResumeRev","productized service","ATS-optimized resume rewrite with a 48h turnaround.","Job seekers (mid-career, tech, new grads)","Resumes get auto-rejected; no replies","More interviews; concrete keyword/format fixes","Stripe/Gumroad, tiered rewrite","$79-129","2-3h","Reddit r/resumes, LinkedIn, Upwork",9,7,9,7,4,6),
 (4,"InMakeover","productized service","LinkedIn profile rewrite that turns views into DMs.","Founders, sales reps, consultants, job seekers","Profile gets views but no inbound","Pipeline & inbound leads from a better profile","Stripe Payment Link","$99","2-3h","LinkedIn DMs, founder communities",8,7,9,6,5,5),
 (5,"GBP Launch","productized service","Done-for-you Google Business Profile setup & optimization.","New local businesses, restaurants, trades","Missing/empty GBP = invisible on Maps","Immediate local visibility & calls","Stripe Payment Link","$129","2-4h","Cold email new businesses, Chamber groups",9,7,8,6,7,5),
 (6,"ColdSeq","productized service","5 ready-to-send cold email sequences written for your offer.","B2B founders, agencies, SDRs","Their cold email gets ignored","Booked meetings from proven frameworks","Gumroad/Stripe","$99","2-3h","Indie Hackers, B2B Slack/Discord",8,7,8,6,5,5),
 (7,"NotionFit","productized service","Custom Notion workspace setup for a freelancer/small team.","Freelancers, coaches, 2-5 person teams","Tool chaos; nothing in one place","Calm, organized ops without DIY hours","Stripe/Calendly","$149","3-5h","Notion subreddit, Twitter, Upwork",7,7,8,5,6,5),
 (8,"ShowNotes","productized service","Podcast show notes + timestamps + 3 social clips per episode.","Indie podcasters, B2B podcast teams","Editing/notes eat hours every week","Recurring time savings; more discoverability","Stripe subscription/per-episode","$39/ep or $129/mo","2-3h","Podcast FB groups, Twitter",8,7,8,5,6,5),
 (9,"A11yQuick","productized service","An ADA/accessibility quick-audit with prioritized fixes.","SMB site owners, agencies, e-com","Legal risk + lost users from inaccessible sites","Reduce lawsuit risk; widen audience","Stripe Payment Link","$199","4-6h","Agency partnerships, cold email",7,8,7,5,7,5),
 (10,"PageTeardown","productized service","A 10-min recorded landing-page teardown with 7 fixes.","Indie founders, marketers, course creators","They suspect their page leaks money but can't see why","Specific conversion fixes from an outside eye","Stripe/Gumroad","$49","1-2h","Twitter, Indie Hackers, Reddit",9,6,9,6,5,5),
 (11,"DeckPolish","productized service","Investor/sales pitch-deck cleanup in 72 hours.","Pre-seed founders, sales teams","Ugly/cluttered decks lose credibility","Sharper story = better meetings","Stripe Payment Link","$199","4-6h","Founder Slacks, accelerators",7,7,7,5,6,4),
 (12,"EtsyBoost","productized service","Etsy listing SEO + photo/title/tag optimization.","Etsy sellers","Listings buried; no traffic","More views & sales without ads","Gumroad/Stripe","$79","2-3h","Etsy seller FB groups, Reddit",8,7,8,6,5,6),
 (13,"NewsletterDFY","productized service","Done-for-you newsletter setup (Beehiiv/Buttondown) + first issue.","Coaches, local brands, creators","Want a newsletter but never start","A launched, branded newsletter day one","Stripe/Calendly","$179","3-5h","Creator Twitter, LinkedIn",7,7,7,5,6,4),
 (14,"BotSetup","productized service","No-code FAQ chatbot installed on a small-biz site.","Local services, clinics, SaaS landing","Repetitive customer questions eat time","Capture leads & answer 24/7","Stripe + optional retainer","$149 + $29/mo","3-5h","Cold email, local groups",7,7,6,5,6,5),
 (15,"StaticMove","productized service","Migrate a slow WordPress brochure site to fast static hosting.","Small businesses on bloated WP","Slow, hacked-prone, costly WP hosting","Faster, safer, cheaper site","Stripe Payment Link","$299","6-10h","Cold email, web-dev subreddits",7,8,6,5,6,4),
 (16,"SecondBrain OS","template pack","A Notion 'second brain' template for capture-to-action.","Knowledge workers, students, creators","Notes everywhere, nothing actionable","A system that actually gets used","Gumroad/Lemon Squeezy","$29","6-10h","Notion subreddit, Twitter, TikTok",8,8,7,5,4,6),
 (17,"FreelanceKit","template pack","Contract + invoice + proposal template pack for freelancers.","New & solo freelancers","Unprofessional docs; unpaid invoices","Look pro & get paid faster","Gumroad","$24","5-8h","r/freelance, Upwork communities",8,8,8,5,6,6),
 (18,"ColdSwipe 100","template pack","100 proven cold-email templates by use-case.","Founders, agencies, SDRs","Blank-page paralysis writing outreach","Faster pipeline from swipeable copy","Gumroad/Lemon Squeezy","$39","6-9h","Indie Hackers, LinkedIn, B2B Discord",8,8,8,5,5,6),
 (19,"ATS Resume Kit","template pack","ATS-friendly resume + cover-letter template bundle.","Job seekers","Templates that break in ATS","Clean, parseable docs that pass filters","Gumroad","$19","5-8h","r/jobs, LinkedIn, TikTok",8,7,8,6,4,7),
 (20,"PitchTen","template pack","A 10-slide investor pitch-deck template + guidance.","Pre-seed/seed founders","Don't know what slides VCs expect","A VC-ready structure fast","Gumroad/Lemon Squeezy","$29","5-8h","Founder Twitter, accelerators",7,7,8,5,6,5),
 (21,"ContentCal","template pack","30-day social content calendar + 100 hook templates.","Solo creators, SMM managers","Run out of ideas; inconsistent posting","Never stare at a blank calendar","Gumroad","$19","5-8h","Instagram, TikTok, Twitter",8,7,8,5,4,6),
 (22,"SaaSModel","template pack","A SaaS financial model spreadsheet (MRR, churn, runway).","Early SaaS founders","Investors want numbers they can't build","Credible model without a finance hire","Lemon Squeezy/Gumroad","$49","8-12h","Indie Hackers, founder Slacks",7,8,6,5,7,6),
 (23,"WeddingSheet","template pack","Wedding budget + guest + timeline planning spreadsheet.","Engaged couples, planners","Spreadsheet chaos; overspending","Stress-free, on-budget planning","Gumroad/Etsy","$15","5-8h","Pinterest, wedding subreddits, Etsy",8,7,8,6,4,7),
 (24,"MacroMeal","template pack","Meal-prep + macro tracking spreadsheet/Notion.","Gym-goers, busy parents","Tracking food is tedious","Hit goals without an app subscription","Gumroad/Etsy","$14","5-8h","Fitness IG/TikTok, r/fitness",7,7,8,6,4,7),
 (25,"HostBook","template pack","Airbnb host welcome-book + house-manual template.","Airbnb/STR hosts","Repeating the same guest info; bad reviews","Better reviews, fewer messages","Gumroad/Etsy","$19","5-8h","STR host FB groups, Reddit",8,7,8,6,5,6),
 (26,"EtsyOps","template pack","Etsy shop policies + listing + announcement template pack.","New Etsy sellers","Don't know what to write for policies","Launch a credible shop fast","Gumroad/Etsy","$15","4-7h","Etsy seller groups, Pinterest",8,7,8,6,5,6),
 (27,"UGC Kit","template pack","UGC creator media-kit + rate-card + pitch template.","Aspiring UGC creators","Don't know how to pitch brands","Land paid brand deals faster","Gumroad","$19","4-7h","TikTok, IG, creator Discords",8,7,8,6,4,6),
 (28,"ListingFlyer","template pack","Editable real-estate listing flyer + social templates (Canva).","Real-estate agents","Marketing materials look amateur","Polished listings that win sellers","Gumroad/Etsy","$24","5-8h","Agent FB groups, Instagram",7,7,8,5,6,5),
 (29,"LaunchKit","template pack","Course-creator launch checklist + email + page templates.","Course/cohort creators","Launches are chaotic & under-planned","A repeatable launch that converts","Gumroad/Lemon Squeezy","$39","6-9h","Creator Twitter, communities",7,7,7,5,5,5),
 (30,"JobTrackr","template pack","Job-search tracker (Notion + Sheets) with follow-up system.","Active job seekers","Lose track of applications & follow-ups","Organized search = more offers","Gumroad","$12","4-6h","r/jobs, LinkedIn, TikTok",7,6,9,6,5,6),
 (31,"DebtSnow","template pack","Debt-payoff (snowball/avalanche) + budget spreadsheet.","People paying down debt","Don't see a path out of debt","A clear, motivating payoff plan","Gumroad/Etsy","$15","5-8h","r/personalfinance, Pinterest",7,7,8,6,4,7),
 (32,"OnePager","template pack","Startup one-pager + data-room checklist template.","Founders raising or partnering","No crisp summary to send investors","Send a pro one-pager in minutes","Gumroad","$19","4-6h","Founder communities",6,6,8,4,6,4),
 (33,"RepurposeSys","template pack","Content repurposing system: 1 video -> 10 posts template.","Creators, marketers","Content dies after one post","10x reach from existing content","Gumroad/Lemon Squeezy","$24","5-8h","Twitter, LinkedIn, YouTube",7,7,8,5,4,5),
 (34,"SheetCRM","template pack","A lightweight CRM in Google Sheets for solopreneurs.","Solopreneurs, freelancers","Don't want to pay for/learn a CRM","Track deals without SaaS bloat","Gumroad","$19","6-9h","r/smallbusiness, Twitter",7,6,7,5,5,5),
 (35,"Goal90","template pack","90-day goal + habit planner (printable + Notion).","Self-improvement crowd","Goals fizzle by week two","Structure & accountability","Gumroad/Etsy","$12","4-7h","Pinterest, IG, TikTok",7,6,8,6,3,6),
 (36,"RateRight","calculator","Freelance rate calculator: target income -> hourly/project rate.","Freelancers, consultants","Underpricing; guessing rates","Confidence to charge correctly","Affiliate (invoicing) + Pro PDF","Free + $9 PDF","3-5h","r/freelance, SEO, Twitter",6,5,8,5,5,7),
 (37,"SaaSMetrics","calculator","SaaS pricing & MRR/churn/LTV calculator.","SaaS founders, PMMs","Pricing & metrics confusion","Clarity on pricing & growth math","Affiliate + consult lead magnet","Free + lead capture","4-6h","Indie Hackers, SEO",6,6,7,5,6,7),
 (38,"PayoffPlus","calculator","Mortgage extra-payment payoff calculator with savings chart.","Homeowners","Don't see impact of extra payments","See years/interest saved instantly","Affiliate (refi/lenders) + ads","Free (affiliate)","4-6h","SEO long-tail, Pinterest",6,7,7,4,4,9),
 (39,"TakeHome","calculator","Contractor/freelancer take-home & tax estimator (per region).","Contractors, gig workers","Surprised by taxes; bad pricing","Price right & avoid tax shocks","Affiliate (accounting) + Pro","Free + $9","6-9h","SEO, r/freelance",6,6,6,4,5,8),
 (40,"ReadCheck","micro-tool","Readability + reading-time + word-count checker.","Writers, students, marketers","Copy too dense; unsure of grade level","Clearer copy that converts/passes","Affiliate (writing tools) + Pro","Free + $9","3-4h","SEO, writing communities",4,4,9,4,3,7),
 (41,"PixSqueeze","micro-tool","Client-side image compressor (no upload, private).","Bloggers, devs, store owners","Big images slow sites; privacy worries","Fast, private compression for free","Ads/affiliate + Pro batch unlock","Free + $9","4-6h","SEO, dev communities",4,5,7,4,3,8),
 (42,"PDFKit","micro-tool","Client-side PDF merge/split/reorder (private, in-browser).","Students, admins, freelancers","Online PDF tools upload your files","Private PDF edits, no signup","Pro unlock + affiliate","Free + $12","6-9h","SEO, r/productivity",5,6,6,4,3,9),
 (43,"BrandPalette","generator","Brand color-palette generator + export tokens.","Designers, indie founders","Picking cohesive colors is hard","Ready palettes + CSS/Tailwind tokens","Affiliate (design) + Pro packs","Free + $9","4-6h","Design Twitter, SEO",4,4,8,4,3,6),
 (44,"MetaPreview","micro-tool","Meta/OpenGraph tag generator + social preview.","Devs, marketers, SEOs","Bad link previews; missing tags","Perfect previews & tags in 1 min","Affiliate (SEO tools) + Pro","Free + $9","4-6h","SEO, dev communities",5,5,8,4,4,8),
 (45,"UTM+QR","micro-tool","UTM link builder + QR code generator.","Marketers, event organizers","Messy tracking links; manual QR","Clean tracking + instant QR","Affiliate + Pro (branded QR)","Free + $9","3-5h","Marketing communities, SEO",5,5,8,4,4,7),
 (46,"InvoiceNow","micro-tool","Client-side invoice generator -> PDF, no signup.","Freelancers, small biz","Invoicing apps want subscriptions","Send a clean invoice free, instantly","Affiliate (accounting) + Pro template","Free + $9","5-8h","r/freelance, SEO",6,6,6,5,4,8),
 (47,"PolicyGen","generator","Privacy policy + terms generator (template-based).","Indie devs, small sites, app makers","App stores/ads require policies","Compliant-ish docs in minutes","Pro (custom clauses) + affiliate","Free + $19","5-8h","Dev communities, SEO",6,6,7,5,5,8),
 (48,"SocialSize","micro-tool","Image resizer/cropper for every social format.","Creators, SMMs","Wrong sizes; manual cropping","One upload, all sizes, in-browser","Pro batch + affiliate","Free + $9","4-6h","Creator communities, SEO",4,5,8,4,3,7),
 (49,"PassForge","micro-tool","Privacy-first password & passphrase generator.","Everyone, security-minded","Weak/reused passwords","Strong, memorable, private generation","Affiliate (password managers)","Free (affiliate)","2-4h","SEO, security communities",4,5,9,4,2,7),
 (50,"DataFlip","micro-tool","CSV<->JSON converter + cleaner (client-side).","Devs, analysts, ops","Tedious format conversion","Quick private conversion, no upload","Pro (bulk/schemas) + affiliate","Free + $9","4-6h","Dev communities, SEO",4,5,8,4,3,8),
 (51,"MDFormat","micro-tool","Markdown -> clean HTML / formatter / TOC.","Writers, devs, bloggers","Messy markdown to publish","Clean output & TOC instantly","Affiliate + Pro","Free + $9","3-5h","Dev/writing communities, SEO",4,4,9,4,3,6),
 (52,"FocusTime","micro-tool","Pomodoro/focus timer with stats & ambient themes.","Students, remote workers","Procrastination; bland timers","Nicer focus sessions; Pro themes","Pro themes (Gumroad) + affiliate","Free + $9","4-6h","r/productivity, TikTok",4,4,8,4,2,6),
 (53,"SubnetPro","calculator","Subnet / CIDR / IP calculator for network admins.","Network engineers, students","Manual subnetting is error-prone","Fast, correct subnetting","Affiliate (cert courses) + Pro","Free + $9","4-6h","Networking communities, SEO",4,5,7,4,5,8),
 (54,"CronWiz","micro-tool","Cron expression builder + plain-English explainer.","Developers, DevOps","Cron syntax is cryptic","Get cron right the first time","Affiliate (hosting/dev tools)","Free (affiliate)","3-5h","Dev communities, SEO",4,4,8,4,5,8),
 (55,"FXConvert","calculator","Currency & unit converter (free ECB daily feed).","Travelers, shoppers, freelancers","Need quick conversions w/ context","Fast conversions + rate history","Affiliate (cards/transfers) + ads","Free (affiliate)","4-6h","SEO long-tail",4,5,7,4,2,8),
 (56,"CSSForge","generator","CSS gradient / box-shadow / glassmorphism generator.","Front-end devs, designers","Tweaking CSS by hand is slow","Copy-paste CSS instantly","Affiliate (UI kits) + Pro packs","Free + $9","4-6h","Dev Twitter, SEO",4,4,8,4,3,7),
 (57,"FaviconFast","generator","Favicon generator from text/emoji/upload (all sizes).","Indie devs, makers","Generating favicons is fiddly","All formats + manifest in 1 click","Affiliate + Pro","Free + $9","4-6h","Dev communities, SEO",4,4,8,4,4,7),
 (58,"OGImage","generator","Open Graph social image generator (templated, client-side).","Bloggers, marketers, devs","Ugly/absent share images","Branded share images fast","Pro templates + affiliate","Free + $12","6-9h","Dev/marketing communities, SEO",5,5,6,4,4,7),
 (59,"TDEEpro","calculator","TDEE / macro / calorie calculator with goal modes.","Dieters, gym-goers","Confusing calorie math","Personalized targets instantly","Affiliate (supps/coaching) + Pro PDF","Free + $9","4-6h","SEO huge, fitness IG/TikTok",5,6,8,4,2,9),
 (60,"DueDate","calculator","Ovulation / due-date / pregnancy calculator.","Expecting & trying parents","Confusing date math; anxiety","Clear dates & milestones","Affiliate (baby products) + ads","Free (affiliate)","3-5h","SEO, parenting communities",5,6,8,4,3,9),
 (61,"SplitEasy","calculator","Tip & bill-splitter with tax/round options.","Diners, groups, travelers","Awkward bill math","Fair splits in seconds","Ads/affiliate","Free","2-3h","SEO, app-alternative searches",3,3,9,3,2,7),
 (62,"SleepCycle","calculator","Bedtime/wake sleep-cycle calculator.","Tired people, shift workers","Wake up groggy","Wake between cycles, feel rested","Affiliate (sleep products) + ads","Free (affiliate)","2-4h","SEO, Pinterest",4,5,9,4,3,8),
 (63,"FIREcalc","calculator","Savings-rate / FIRE / coast-FIRE calculator.","FIRE & PF community","When can I retire? unclear","See your FIRE date & levers","Affiliate (brokerages/PF) + Pro","Free + $9","5-7h","r/financialindependence, SEO",5,6,7,4,4,8),
 (64,"QuoteEstimator","calculator","Project quote/estimate builder for freelancers/agencies.","Freelancers, small agencies","Inconsistent, slow quoting","Confident quotes fast","Pro (branded PDF) + affiliate","Free + $19","5-8h","r/freelance, agency communities",6,6,7,5,5,7),
 (65,"ROAScalc","calculator","Ad ROAS / break-even / CAC calculator.","E-com & performance marketers","Unsure if ads are profitable","Know break-even before scaling","Affiliate (ad tools) + Pro","Free + $9","4-6h","Ecom/marketing communities, SEO",5,6,7,4,5,7),
 (66,"AIToolsForLawyers","directory","Curated directory of AI tools for legal professionals.","Lawyers, paralegals, legal ops","Hard to find vetted, safe AI tools","Save research time; avoid risky tools","Featured listings + affiliate","Free; $99 featured","8-12h","Legal communities, SEO, LinkedIn",5,7,6,4,7,8),
 (67,"RemoteFirst","directory","Directory of genuinely remote-first companies hiring.","Remote job seekers","Fake-remote listings everywhere","Find real remote employers fast","Job-post fees + affiliate","Free; $79 post","8-12h","r/remotework, SEO, LinkedIn",5,7,6,4,5,8),
 (68,"NoCodeStack","directory","No-code tools directory by use-case with comparisons.","Non-technical founders","Tool overwhelm; which to pick","Pick the right no-code tool fast","Affiliate + featured listings","Free; affiliate","8-12h","No-code communities, SEO",5,6,6,4,5,8),
 (69,"BoilerHub","directory","Directory of SaaS boilerplates & starter kits (affiliate).","Indie devs, founders","Which boilerplate to buy?","Save weeks; pick the right kit","Affiliate (boilerplate sales)","Free; affiliate","6-10h","Indie Hackers, dev Twitter, SEO",5,7,7,4,6,7),
 (70,"FreeTierDev","directory","Directory of free-tier dev tools & limits, compared.","Indie devs, students","Hard to compare free limits","Build for $0 with the right tools","Affiliate + sponsor slot","Free; sponsor","6-10h","Dev communities, SEO",5,6,7,4,5,8),
 (71,"MakerStack","directory","Directory of the tools indie makers actually use.","Indie hackers, solopreneurs","Decision fatigue on tooling","Copy proven stacks","Affiliate + featured","Free; affiliate","6-10h","Indie Hackers, Twitter",4,6,7,4,5,7),
 (72,"SponsorList","directory","Directory of newsletters open to sponsorship by niche.","B2B/DTC advertisers","Hard to find newsletters to sponsor","Find sponsorship inventory fast","Featured listings + finder fee","Free; $99 featured","8-12h","Marketing communities, SEO",5,7,6,4,7,7),
 (73,"GrantFinder","directory","Directory of grants/accelerators for a region/niche.","Founders, nonprofits","Grants are scattered & hard to find","Find funding you qualify for","Featured + premium filter","Free; $9 premium","8-12h","Founder/nonprofit communities, SEO",5,6,6,4,6,8),
 (74,"StartupPerks","directory","Directory of startup discounts/credits (affiliate).","Early founders","Missing free credits & deals","Save thousands on tools","Affiliate + featured","Free; affiliate","6-10h","Indie Hackers, founder Slacks",5,6,7,4,5,7),
 (75,"CoworkCity","directory","Best coworking spaces in a target city, filterable.","Remote workers, nomads","Finding good coworking is tedious","Pick the right space fast","Featured listings + affiliate","Free; $49 featured","6-10h","Nomad communities, local SEO",4,6,7,4,5,8),
 (76,"NicheJobs","directory","Curated job board for a specific niche (e.g., AI/ML).","Specialised seekers & employers","Generic boards are noisy","Signal over noise for niche roles","Job-post fees","Free; $99 post","8-12h","Niche communities, SEO, LinkedIn",5,7,6,4,5,8),
 (77,"GuestMatch","directory","Podcast guest<->host matchmaking directory.","Podcasters & potential guests","Booking good guests/spots is hard","Faster, better bookings","Featured profiles + intro fee","Free; $29 featured","8-12h","Podcast communities, Twitter",4,6,6,4,6,6),
 (78,"OpenAlt","directory","'X but open-source' alternatives directory.","Privacy-minded users, devs","Want OSS alternatives to SaaS","Find trustworthy OSS fast","Affiliate (hosting) + sponsor","Free; sponsor","6-10h","r/selfhosted, HN, SEO",5,6,7,4,4,9),
 (79,"ExtPicks","directory","Curated Chrome-extension directory by workflow.","Power users, professionals","Store is noisy; trust issues","Vetted extensions by job","Affiliate + featured","Free; affiliate","6-9h","Productivity communities, SEO",4,5,7,4,4,8),
 (80,"HustleDB","directory","Searchable database of side-hustle ideas + playbooks.","Aspiring side-hustlers","Idea overwhelm; no playbooks","Pick & start a hustle fast","Premium playbooks + affiliate","Free; $19 premium","8-12h","r/sidehustle, TikTok, SEO",5,6,6,4,4,8),
 (81,"SolarQuote","lead-gen site","Local solar-quote lead capture by region.","Homeowners considering solar","Confusing, pushy solar sales","Easy, no-pressure quote comparison","Sell leads to installers","~$20-60/lead","6-9h","Local SEO + FB ads test",6,8,7,4,5,7),
 (82,"RoofLeads","lead-gen site","Roofing/HVAC quote lead-gen for a metro.","Homeowners needing repairs","Hard to find trusted local pros","Fast quotes from vetted pros","Sell leads to contractors","~$25-75/lead","6-9h","Local SEO + FB ads",6,8,7,4,5,7),
 (83,"RefiCheck","lead-gen site","Mortgage refinance interest capture + match.","Homeowners","Unsure if refinancing is worth it","See if refi saves money, get matched","Affiliate/referral to lenders","Per-lead referral","6-9h","SEO + content",5,8,7,4,4,7),
 (84,"RestaurantSite","lead-gen site","Web-design lead-gen targeting restaurants without sites.","Restaurants w/ poor/no website","Losing orders to better-presented rivals","More covers from a real site","Feeds your own web-design service","$0 (feeds service)","5-8h","Local outreach + walk-ins",6,7,7,5,6,5),
 (85,"InsureCompare","lead-gen site","Insurance comparison lead capture (niche).","Shoppers for a specific insurance","Comparing policies is painful","Find cheaper coverage fast","Affiliate/lead sale to brokers","Per-lead","6-9h","SEO + paid test",5,8,6,4,4,7),
 (86,"MoveQuote","lead-gen site","Moving-company quote lead-gen for a metro.","People relocating","Getting movers' quotes is a hassle","Compare movers in one form","Sell leads to movers","~$15-40/lead","5-8h","Local SEO + FB ads",5,7,7,4,5,7),
 (87,"TrialDesk","lead-gen site","Gym/PT free-trial signup lead-gen (local).","Local gyms & trainers","Empty trial pipeline","Steady trial bookings","Per-lead or monthly to gym","~$10-25/lead","5-8h","Local FB/IG, walk-ins",5,7,7,5,5,5),
 (88,"VendorInq","lead-gen site","Wedding-vendor inquiry lead-gen (photographers etc.).","Wedding vendors","Inconsistent inquiry flow","Booked weddings from steady leads","Per-lead to vendors","~$15-40/lead","5-8h","Pinterest, local SEO",5,7,7,4,5,6),
 (89,"NameSpark","generator","Business name + domain-idea generator (wordlist engine).","Founders, side-hustlers","Naming paralysis; domains taken","Fast names with open domains","Affiliate (registrars) + Pro packs","Free (affiliate)","4-6h","SEO, founder communities",5,5,8,4,3,8),
 (90,"SloganGen","generator","Slogan / tagline generator by industry (template engine).","Small businesses, founders","Can't phrase a punchy tagline","Tagline options in seconds","Affiliate + Pro pack","Free + $9","3-5h","SEO, small-biz communities",4,4,8,4,3,7),
 (91,"BulletBoost","generator","Resume bullet-point generator (action-verb + metric templates).","Job seekers","Weak, vague resume bullets","Strong, quantified bullets fast","Pro pack + resume-service upsell","Free + $9","4-6h","r/resumes, LinkedIn, SEO",6,6,8,5,4,7),
 (92,"BioCraft","generator","Social-media bio generator by niche/platform.","Creators, professionals","Bland, generic bios","On-brand bios that convert follows","Pro pack + affiliate","Free + $9","3-5h","TikTok, IG, SEO",5,5,8,4,3,6),
 (93,"HashSet","generator","Niche hashtag set generator + reach tiers.","Creators, SMMs","Guessing hashtags; poor reach","Curated hashtag sets per niche","Pro packs + affiliate","Free + $9","4-6h","IG/TikTok creator communities",4,4,8,4,3,6),
 (94,"PromptVault","generator","Searchable AI-prompt template library by role/task.","AI users, marketers, devs","Prompts are scattered; results vary","Proven prompts that work","Pro pack (Gumroad) + affiliate","Free + $19","6-9h","Twitter, AI communities, SEO",6,6,7,5,3,7),
 (95,"UserHandle","generator","Username/handle idea generator + availability hints.","Creators, gamers, founders","Desired handle taken everywhere","Available handle ideas fast","Affiliate + Pro","Free + $9","3-5h","Creator/gaming communities, SEO",4,4,8,4,3,6),
 (96,"MetaDesc","generator","Meta-description / SEO snippet generator (templated).","Bloggers, SEOs, small biz","Weak SERP snippets; low CTR","Click-worthy snippets fast","Affiliate (SEO tools) + Pro","Free + $9","3-5h","SEO communities",4,5,8,4,4,7),
 (97,"ColdGen","generator","Cold-email generator: fill inputs -> ready sequence (templates).","Founders, SDRs, agencies","Writing cold email from scratch","Ready-to-send sequences fast","Pro pack + service upsell","Free + $19","5-8h","B2B communities, LinkedIn",6,6,7,5,4,6),
 (98,"InboxZero Waitlist","landing page","Pre-launch waitlist for an inbox-triage tool for founders.","Overwhelmed founders/execs","Email overload kills focus","Reclaim hours every week","Waitlist -> paid beta (Stripe)","Free now; ~$15/mo later","2-4h","Founder Twitter, Indie Hackers",4,7,9,4,3,5),
 (99,"7DayLaunch","landing page","Pre-sell page for a 'launch in 7 days' ebook/course.","Aspiring makers","Stuck, never ship","A concrete shipping plan","Pre-order via Stripe/Gumroad","$29 pre-order","2-4h","Indie Hackers, Twitter",6,7,9,5,4,5),
 (100,"BuildInPublic Kit","landing page","Pre-order page for a build-in-public content system.","Creators/founders building in public","Inconsistent BIP content","A system to post consistently","Pre-order via Gumroad","$24 pre-order","2-4h","Twitter, creator communities",5,6,9,5,4,5),
]

def frs(d): return round(d["rs"]*.30 + d["pp"]*.25 + d["eb"]*.20 + d["as_"]*.15 + d["lc"]*.10, 3)

ideas=[]
for t in RAW:
    d=dict(zip(["id","name","type","desc","target","pain","why_pay","monetization","price","build","channel","rs","pp","eb","as_","lc","seo"],t))
    base=CAT[d["type"]].copy(); base.update(d)
    base["frs"]=frs(base)
    base["launch_difficulty"]=11-base["eb"]
    base["competition_level"]=11-base["lc"]
    base["deploy"]="Vercel (static) - primary; Cloudflare Pages / Netlify / GitHub Pages as alternates"
    ideas.append(base)

ideas_sorted=sorted(ideas,key=lambda x:(-x["frs"],-x["rs"],-x["pp"]))
for i,d in enumerate(ideas_sorted,1): d["rank"]=i

# --- Portfolio selection: FRS order WITH category quotas. -------------------
# Rationale (documented in ASSESSMENT.md): pure FRS over-concentrates in
# outreach-dependent offers (services/templates) that all compete for the same
# daily outreach hours, and selects zero compounding SEO assets (tools,
# directories). Quotas keep the FRS discipline inside each category while
# diversifying acquisition channels: outreach cash now + SEO compounding later.
QUOTAS={"productized service":12,"template pack":12,"directory":7,"lead-gen site":6,
        "landing page":3,"calculator":4,"micro-tool":3,"generator":3,"mini SaaS":0}
counts={k:0 for k in QUOTAS}
top50=[]; near=[]
for d in ideas_sorted:
    if counts.get(d["type"],0) < QUOTAS.get(d["type"],0):
        counts[d["type"]]+=1; top50.append(d)
    else:
        near.append(d)
near_misses=[d for d in near if d["frs"]>=top50[-1]["frs"]-0.7][:10]
for i,d in enumerate(top50,1): d["sel_rank"]=i

top10=top50[:10]
top3_fast=sorted(top50,key=lambda x:(x["rs"],x["frs"]),reverse=True)[:3]
static_pool=[d for d in top50 if d["type"] in ("template pack","calculator","micro-tool","directory","generator","landing page")]
top10_static=sorted(static_pool,key=lambda x:(x["eb"],x["frs"]),reverse=True)[:10]
top10_onetemplate=sorted([d for d in top50 if d["type"] in ("productized service","landing page","template pack","directory")],key=lambda x:x["frs"],reverse=True)[:10]
top5_seo=sorted(top50,key=lambda x:(x["seo"],x["frs"]),reverse=True)[:5]
top5_cold=sorted([d for d in top50 if d["type"] in("productized service","lead-gen site")],key=lambda x:(x["as_"],x["rs"],x["frs"]),reverse=True)[:5]
top5_download=sorted([d for d in top50 if d["type"]=="template pack"],key=lambda x:x["frs"],reverse=True)[:5]
top5_service=sorted([d for d in top50 if d["type"]=="productized service"],key=lambda x:x["frs"],reverse=True)[:5]

HERE=os.path.dirname(os.path.abspath(__file__)); REPO=os.path.dirname(HERE)
DOCS=os.path.join(REPO,"docs"); DATA=os.path.join(REPO,"data")
os.makedirs(DOCS,exist_ok=True); os.makedirs(DATA,exist_ok=True)
json.dump(ideas_sorted,open(os.path.join(DATA,"ideas_all100.json"),"w"),indent=2)
json.dump(top50,open(os.path.join(DATA,"ideas_top50.json"),"w"),indent=2)

# ---------------- Markdown ----------------
def row(d, cols): return "| "+" | ".join(str(d[c]) for c in cols)+" |"
def slug_line(label,val): return f"- **{label}:** {val}"
md=[]
md.append("# Startup Factory — Phase 1: Market Assessment\n")
md.append("> Brutally realistic scoring of 100 ideas. **No profit is guaranteed.** Most ideas validate weakly; the point of the factory is to ship fast and kill losers fast. Scores are judgement-based estimates (1–10), not measured data — treat them as a prioritisation lens, not truth.\n")
md.append("## Scoring model\n")
md.append("`Fast Revenue Score = RevenueSpeed×0.30 + ProfitPotential×0.25 + EaseOfBuild×0.20 + AcquisitionSimplicity×0.15 + LowCompetition×0.10`\n")
md.append("All inputs 1–10, higher = better (incl. *Ease of Build* and *Low Competition*). Reported **Launch difficulty = 11 − EaseOfBuild** and **Competition level = 11 − LowCompetition** (higher = harder/more competitive).\n")
avg=round(sum(d['frs'] for d in top50)/50,2)
spread={c:sum(1 for d in top50 if d['type']==c) for c in CAT}
spread={k:v for k,v in spread.items() if v}
md.append(f"**Top-50 average FRS:** {avg}. **Category spread (top 50):** "+", ".join(f"{k} ×{v}" for k,v in sorted(spread.items(),key=lambda x:-x[1]))+".\n")

md.append("\n## 1. The 100 ideas (scored & ranked)\n")
md.append("| # | Name | Type | RS | PP | EB | AS | LC | SEO | **FRS** |")
md.append("|---:|---|---|:-:|:-:|:-:|:-:|:-:|:-:|:-:|")
for d in ideas_sorted:
    md.append(f"| {d['rank']} | {d['name']} | {d['type']} | {d['rs']} | {d['pp']} | {d['eb']} | {d['as_']} | {d['lc']} | {d['seo']} | **{d['frs']}** |")

def mini_table(title, items, note=""):
    md.append(f"\n## {title}\n")
    if note: md.append(note+"\n")
    md.append("| Rank | Name | Type | One-liner | FRS |")
    md.append("|---:|---|---|---|:-:|")
    for i,d in enumerate(items,1):
        md.append(f"| {i} | **{d['name']}** | {d['type']} | {d['desc']} | {d['frs']} |")

mini_table("2. Top 50 — selected & ranked", top50,
  "Selected by FRS **within category quotas** (services 12, template packs 12, directories 7, lead-gen 6, landing 3, calculators 4, micro-tools 3, generators 3). "
  "Why not pure FRS? Pure FRS picks ~35 outreach-dependent offers that all compete for the same daily outreach hours and zero compounding SEO assets. "
  "Quotas keep FRS discipline inside each category while diversifying acquisition: outreach cash now + SEO assets that compound later. "
  "Near-misses cut by quota: " + ", ".join(f"{d['name']} ({d['frs']})" for d in near_misses) + ".")
mini_table("3. Top 10 overall priorities", top10)
mini_table("4. Top 3 fastest-revenue bets (best shot at money in 14 days)", top3_fast,
  "Fastest cash comes from selling a clear outcome to an identifiable buyer via outreach — not from hoping SEO traffic shows up. These three can realistically take money inside two weeks **if you actually do the outreach.**")
mini_table("5. Top 10 launchable as static sites (almost no backend)", top10_static)
mini_table("6. Top 10 buildable from ONE reusable template", top10_onetemplate)
mini_table("7. Top 5 best for SEO (compounding, slower)", top5_seo)
mini_table("8. Top 5 best for cold outreach", top5_cold)
mini_table("9. Top 5 best for template/download monetization", top5_download)
mini_table("10. Top 5 best for productized-service monetization", top5_service)

md.append("\n## 11. Full 27-field profile — each of the top 50\n")
FIELDS=[("Type","type"),("Description","desc"),("Target customer","target"),("Pain point","pain"),
 ("Why they pay","why_pay"),("Fastest monetization","monetization"),("Price","price"),
 ("MVP features only","mvp"),("Skip for now","skip"),("Free/OSS stack","stack"),
 ("Deployment","deploy"),("Data/storage","storage"),("Payment","pay"),("Est. build time","build"),
 ("Launch difficulty (1-10)","launch_difficulty"),("Revenue speed (1-10)","rs"),
 ("Profit potential (1-10)","pp"),("Competition level (1-10)","competition_level"),
 ("Acquisition simplicity (1-10)","as_"),("SEO potential (1-10)","seo"),
 ("Fast Revenue Score","frs"),("Best acquisition channel","channel"),
 ("First 10 customers","first10"),("14-day validation test","test"),
 ("Biggest risk","risk"),("Kill criteria","kill")]
for d in top50:
    md.append(f"\n### {d['rank']}. {d['name']}  ·  FRS {d['frs']}\n")
    for label,key in FIELDS:
        md.append(slug_line(label,d[key]))
open(os.path.join(DOCS,"ASSESSMENT.md"),"w").write("\n".join(md))

# ---------------- Excel ----------------
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb=Workbook()
    INK="111111"; ACCENT="2C5F4F"; PAPER="F4F1EA"; HEADTXT="FFFFFF"
    head_fill=PatternFill("solid",fgColor=ACCENT); head_font=Font(bold=True,color=HEADTXT,size=11)
    title_font=Font(bold=True,size=15,color=INK); thin=Side(style="thin",color="DDD7CC")
    border=Border(left=thin,right=thin,top=thin,bottom=thin)
    wrap=Alignment(wrap_text=True,vertical="top"); center=Alignment(horizontal="center",vertical="center")

    def style_header(ws,ncol,r=1):
        for c in range(1,ncol+1):
            cell=ws.cell(row=r,column=c); cell.fill=head_fill; cell.font=head_font
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border
        ws.freeze_panes=ws.cell(row=r+1,column=1)

    # Tab 1: All 100 with live formula
    ws=wb.active; ws.title="All 100 (scored)"
    hdr=["Rank","ID","Name","Type","Description","RS","PP","EB","AS","LC","SEO","FRS (formula)"]
    ws.append(hdr); style_header(ws,len(hdr))
    for d in ideas_sorted:
        r=ws.max_row+1
        ws.append([d["rank"],d["id"],d["name"],d["type"],d["desc"],d["rs"],d["pp"],d["eb"],d["as_"],d["lc"],d["seo"],None])
        ws.cell(row=r,column=12).value=f"=F{r}*0.3+G{r}*0.25+H{r}*0.2+I{r}*0.15+J{r}*0.1"
        ws.cell(row=r,column=12).font=Font(bold=True)
    widths=[6,5,18,20,52,5,5,5,5,5,5,12]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    for r in range(2,ws.max_row+1):
        ws.cell(row=r,column=5).alignment=wrap
        for c in range(6,13): ws.cell(row=r,column=c).alignment=center

    # Tab 2: Top 50 ranked
    ws2=wb.create_sheet("Top 50 ranked")
    ws2.append(hdr[:-1]+["FRS"]); style_header(ws2,len(hdr))
    for d in top50:
        ws2.append([d["rank"],d["id"],d["name"],d["type"],d["desc"],d["rs"],d["pp"],d["eb"],d["as_"],d["lc"],d["seo"],d["frs"]])
    for i,w in enumerate(widths,1): ws2.column_dimensions[get_column_letter(i)].width=w
    for r in range(2,ws2.max_row+1):
        ws2.cell(row=r,column=5).alignment=wrap
        for c in range(6,13): ws2.cell(row=r,column=c).alignment=center

    # Tab 3: Sub-rankings
    ws3=wb.create_sheet("Sub-rankings")
    def block(ws,title,items,startcol):
        ws.cell(row=1,column=startcol).value=title; ws.cell(row=1,column=startcol).font=Font(bold=True,color=HEADTXT)
        ws.cell(row=1,column=startcol).fill=head_fill
        ws.cell(row=1,column=startcol+1).fill=head_fill
        ws.cell(row=1,column=startcol+1).value="FRS"; ws.cell(row=1,column=startcol+1).font=Font(bold=True,color=HEADTXT)
        for i,d in enumerate(items,1):
            ws.cell(row=1+i,column=startcol).value=f"{i}. {d['name']}"
            ws.cell(row=1+i,column=startcol+1).value=d["frs"]
        ws.column_dimensions[get_column_letter(startcol)].width=26
        ws.column_dimensions[get_column_letter(startcol+1)].width=7
    blocks=[("Top 10 overall",top10),("Top 3 fastest revenue",top3_fast),("Top 10 static",top10_static),
            ("Top 10 one-template",top10_onetemplate),("Top 5 SEO",top5_seo),("Top 5 cold outreach",top5_cold),
            ("Top 5 download",top5_download),("Top 5 service",top5_service)]
    col=1
    for title,items in blocks:
        block(ws3,title,items,col); col+=3

    # Tab 4: Per-project 27-field detail (top 50)
    ws4=wb.create_sheet("Top 50 detail (27 fields)")
    detail_cols=["Rank","Name"]+[lbl for lbl,_ in FIELDS]
    ws4.append(detail_cols); style_header(ws4,len(detail_cols))
    for d in top50:
        ws4.append([d["rank"],d["name"]]+[d[key] for _,key in FIELDS])
    ws4.column_dimensions["A"].width=6; ws4.column_dimensions["B"].width=18
    for i in range(3,len(detail_cols)+1):
        ws4.column_dimensions[get_column_letter(i)].width=30
    for r in range(2,ws4.max_row+1):
        for c in range(1,len(detail_cols)+1): ws4.cell(row=r,column=c).alignment=wrap

    # Tab 5: Scoring model
    ws5=wb.create_sheet("Scoring model")
    ws5["A1"]="Fast Revenue Score — weighting"; ws5["A1"].font=title_font
    rows=[["Input","Weight","Meaning (1-10, higher=better)"],
          ["Revenue Speed",0.30,"How fast the first dollar can realistically arrive"],
          ["Profit Potential",0.25,"Margin & ceiling once it works"],
          ["Ease of Build",0.20,"How quickly a real MVP ships (high=easy)"],
          ["Acquisition Simplicity",0.15,"How easy to reach & convert first buyers"],
          ["Low Competition",0.10,"Headroom vs incumbents (high=less crowded)"],
          ["","",""],
          ["Reported Launch difficulty","","11 - Ease of Build"],
          ["Reported Competition level","","11 - Low Competition"]]
    for r in rows: ws5.append(r)
    style_header(ws5,3,r=3)
    ws5.column_dimensions["A"].width=26; ws5.column_dimensions["B"].width=10; ws5.column_dimensions["C"].width=60
    for rr in range(4,ws5.max_row+1): ws5.cell(row=rr,column=3).alignment=wrap

    wb.save(os.path.join(DOCS,"scoring.xlsx"))
    print("xlsx OK")
except Exception as e:
    print("xlsx SKIPPED:",e)

print("Top 3 by FRS:", [(d['name'],d['frs']) for d in top50[:3]])
print("Top 3 fastest-revenue:", [d['name'] for d in top3_fast])
print("avg FRS top50:",avg,"| spread:",spread)
print("DONE -> docs/ASSESSMENT.md, docs/scoring.xlsx, data/*.json")
